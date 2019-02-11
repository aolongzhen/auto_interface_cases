__author__ = 'zz'
#把测试数据从Excel里面获取出来
from openpyxl import  load_workbook

class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
    #读取未注册手机号
    def no_reg_tel(self):
        wb=load_workbook(self.file_path)
        sheet=wb['init']
        no_reg_tel=sheet.cell(1,2).value
        return no_reg_tel#从Excel里面获取到未注册的手机号

    def read_data(self,mode,case_list):
        no_reg_tel=self.no_reg_tel()
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        test_data=[]#存储所有行的数据
        #方法一：
        # if mode=='1':#执行所有的用例 就要加载所有的用例
        #     for i in range(2,sheet.max_row+1):
        #         sub_data={}#存储每一行的数据or存到一个字典里面sub_data={}
        #         sub_data['case_id']=sheet.cell(i,1).value
        #         sub_data['method']=sheet.cell(i,4).value
        #         sub_data['url']=sheet.cell(i,5).value
        #         sub_data['expectresult']=sheet.cell(i,7).value
        #
        #         #对请求参数param进行替换
        #         if sheet.cell(i,6).value.find('${first_tel}')!=-1:
        #            new_param=sheet.cell(i,6).value.replace('${first_tel}',str(no_reg_tel))
        #         else:
        #            new_param=sheet.cell(i,6).value
        #         sub_data['params']=new_param
        #
        #         #对check_sql去进行更新值
        #         if sheet.cell(i,8).value!=None and sheet.cell(i,8).value.find('${no_reg_tel}')!=-1:
        #            new_param=sheet.cell(i,8).value.replace('${no_reg_tel}',str(no_reg_tel))
        #         else:
        #            new_param=sheet.cell(i,8).value
        #         sub_data['check_sql']=new_param
        #         test_data.append(sub_data)
        # elif mode=='0':
        #     for i in case_list:
        #         sub_data={}#存储每一行的数据or存到一个字典里面sub_data={}
        #         sub_data['case_id']=sheet.cell(i+1,1).value
        #         sub_data['method']=sheet.cell(i+1,4).value
        #         sub_data['url']=sheet.cell(i+1,5).value
        #         sub_data['expect_result']=sheet.cell(i+1,7).value
        #
        #         #对请求参数param进行替换
        #         if sheet.cell(i+1,6).value.find('${no_reg_tel}')!=-1:
        #            new_param=sheet.cell(i+1,6).value.replace('${no_reg_tel}',str(no_reg_tel))
        #         else:
        #            new_param=sheet.cell(i+1,6).value
        #         sub_data['params']=new_param
        #
        #         #对check_sql去进行更新值
        #         if sheet.cell(i+1,8).value.find('${fno_reg_tel}')!=-1:
        #            new_param=sheet.cell(i+1,8).value.replace('${no_reg_tel}',str(no_reg_tel))
        #         else:
        #            new_param=sheet.cell(i+1,8).value
        #         sub_data['check_sql']=new_param
        #         test_data.append(sub_data)
        #方法二：
        for i in range(2,sheet.max_row+1):
            sub_data={}#存储每一行的数据or存到一个字典里面sub_data={}
            sub_data['case_id']=sheet.cell(i,1).value
            sub_data['method']=sheet.cell(i,4).value
            sub_data['url']=sheet.cell(i,5).value
            #针对充值的时候，对期望值的手机号做一个改动
            if sheet.cell(i,7).value.find('${no_reg_tel}')!=-1:
                 new_param=sheet.cell(i,7).value.replace('${no_reg_tel}',str(no_reg_tel))
            else:
                new_param=sheet.cell(i,7).value
            sub_data['expect_result']=new_param

            #对请求参数param进行替换
            if sheet.cell(i,6).value.find('${no_reg_tel}')!=-1:
                new_param=sheet.cell(i,6).value.replace('${no_reg_tel}',str(no_reg_tel))
            else:
                new_param=sheet.cell(i,6).value
            sub_data['params']=new_param

            #对check_sql去进行更新值
            if sheet.cell(i,8).value!=None and sheet.cell(i,8).value.find('${no_reg_tel}')!=-1:
                new_param=sheet.cell(i,8).value.replace('${no_reg_tel}',str(no_reg_tel))
            else:
                new_param=sheet.cell(i,8).value
            sub_data['check_sql']=new_param
            test_data.append(sub_data)#所有的数据都存在test_data里面
        if mode=='1':
            final_data=test_data#返回所有的数据
        else:
            final_data=[]
            for item in test_data:
                if item['case_id'] in case_list:
                    final_data.append(item)
        self.update_tel(str(int(no_reg_tel)+1))#更新Excel中的手机号码
        return final_data

    def update_tel(self,new_tel):
        wb=load_workbook(self.file_path)
        sheet=wb['init']
        sheet.cell(1,2).value=new_tel
        wb.save(self.file_path)

    def write_data(self,row,mode,actually,result):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        if mode==1:
            sheet.cell(row,9).value=actually
            sheet.cell(row,10).value=result
        elif mode==2:
            sheet.cell(row,11).value=actually
            sheet.cell(row,12).value=result
        wb.save(self.file_path)

if __name__ == '__main__':
    from conf import project_path
    test_data=DoExcel(project_path.test_data_path,'test_data').read_data('0',[4,5])
    print(test_data)